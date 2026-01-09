import openpyxl
from openpyxl.styles import PatternFill
import io

# --- НАСТРОЙКИ СТОЛБЦОВ ---
COL_ID = 1  # A: ID
COL_KVARTAL = 3  # C: Квартал
COL_NAME = 5  # E: Название
COL_QTY_PLAN = 9  # I: План (шт)
COL_SUM_PLAN = 25  # Y: Сумма План (с НДС)
COL_CATEGORY = 49  # AW: Категория

# Скрытый столбец для хранения ФАКТА (цифр)
COL_QTY_FACT = 60

FILE_NAME = 'table.xlsx'

# === ЦВЕТА ===
POSSIBLE_GREENS = [
    'FF00B050', '00B050',
    'FF92D050', '92D050',
    'FF00FF00', '00FF00',
    'FF008000', '008000',
    'FFC6EFCE',
]

IGNORE_COLORS = [
    '00000000', 'FFFFFFFF', '64', '000000', 'None', None
]

# Цвета для заливки ботом
COLOR_FULL = 'FF00B050'  # Зеленый
COLOR_PARTIAL = 'FFFFFF00'  # Желтый
FILL_FULL = PatternFill(start_color=COLOR_FULL, end_color=COLOR_FULL, fill_type='solid')
FILL_PARTIAL = PatternFill(start_color=COLOR_PARTIAL, end_color=COLOR_PARTIAL, fill_type='solid')


def load_data():
    wb = openpyxl.load_workbook(FILE_NAME)
    ws = wb.active
    # Инициализация заголовка для факта
    if not ws.cell(row=1, column=COL_QTY_FACT).value:
        ws.cell(row=1, column=COL_QTY_FACT).value = "Кол-во ФАКТ"
    return wb, ws


def is_row_green(row):
    """Проверка цвета строки (учитываем ручную заливку)"""
    # Проверяем ID и Сумму
    cells_to_check = [row[0], row[COL_SUM_PLAN - 1]]

    for cell in cells_to_check:
        try:
            color = cell.fill.start_color.index
            color_str = str(color).upper().strip()

            if color_str in IGNORE_COLORS: continue
            if color_str in POSSIBLE_GREENS: return True
        except:
            continue
    return False


# --- ФУНКЦИЯ, КОТОРУЮ Я ЗАБЫЛ В ПРОШЛЫЙ РАЗ ---
def get_analytics(filter_type, value=None):
    """Сбор детальной статистики для кнопок"""
    wb, ws = load_data()

    total_sum_plan = 0
    total_sum_fact = 0
    total_qty_plan = 0
    total_qty_fact = 0
    items_found = []

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            # Читаем основные данные
            row_cat = str(row[COL_CATEGORY - 1].value or "")
            row_kv = str(row[COL_KVARTAL - 1].value or "")

            qty_plan = float(row[COL_QTY_PLAN - 1].value or 0)
            sum_plan = float(row[COL_SUM_PLAN - 1].value or 0)

            # Читаем факт (УМНАЯ ЛОГИКА)
            qty_fact = float(ws.cell(row=idx, column=COL_QTY_FACT).value or 0)

            # Если цифр нет, но строка ЗЕЛЕНАЯ -> считаем Выполненным (Факт = План)
            is_green_visually = is_row_green(row)
            if qty_fact == 0 and is_green_visually:
                qty_fact = qty_plan

                # Считаем деньги факта пропорционально
            price_per_unit = sum_plan / qty_plan if qty_plan > 0 else 0
            sum_fact = qty_fact * price_per_unit
            if sum_fact > sum_plan: sum_fact = sum_plan

            # ФИЛЬТРАЦИЯ
            match = False
            if filter_type == 'total':
                match = True
            elif filter_type == 'quarter' and str(value) == row_kv:
                match = True
            elif filter_type == 'category' and str(value).lower() in row_cat.lower():
                match = True

            if match:
                total_sum_plan += sum_plan
                total_sum_fact += sum_fact
                total_qty_plan += qty_plan
                total_qty_fact += qty_fact

                # Добавляем в список (только первые 50 для скорости)
                if len(items_found) < 50:
                    items_found.append({
                        'name': str(row[COL_NAME - 1].value or "---"),
                        'qty': qty_fact,
                        'is_received': (qty_fact >= qty_plan)
                    })
        except Exception:
            continue

    return {
        'sum_plan': total_sum_plan,
        'sum_fact': total_sum_fact,
        'qty_plan': total_qty_plan,
        'qty_fact': total_qty_fact,
        'items': items_found
    }


def get_item_info(search_term):
    """Поиск товара для приемки"""
    wb, ws = load_data()
    search_term = str(search_term).strip().lower()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        cell_id = str(row[COL_ID - 1].value or "").strip().lower()
        cell_name = str(row[COL_NAME - 1].value or "").strip().lower()

        if search_term == cell_id or (len(search_term) > 3 and search_term in cell_name):
            qty_plan = float(row[COL_QTY_PLAN - 1].value or 0)
            qty_fact = float(ws.cell(row=row_idx, column=COL_QTY_FACT).value or 0)

            if qty_fact == 0 and is_row_green(row):
                qty_fact = qty_plan

            return {
                'found': True,
                'row_idx': row_idx,
                'name': row[COL_NAME - 1].value,
                'id': row[COL_ID - 1].value,
                'plan': qty_plan,
                'fact': qty_fact,
            }
    return {'found': False}


def update_item_qty(row_idx, added_qty):
    """Обновление количества через бота"""
    wb, ws = load_data()

    qty_plan = float(ws.cell(row=row_idx, column=COL_QTY_PLAN).value or 0)
    current_fact = float(ws.cell(row=row_idx, column=COL_QTY_FACT).value or 0)

    new_fact = current_fact + added_qty
    if new_fact < 0: new_fact = 0

    ws.cell(row=row_idx, column=COL_QTY_FACT).value = new_fact

    row_cells = ws[row_idx]
    fill_to_apply = None

    if new_fact >= qty_plan:
        fill_to_apply = FILL_FULL
    elif new_fact > 0:
        fill_to_apply = FILL_PARTIAL
    else:
        fill_to_apply = PatternFill(fill_type=None)

    for cell in row_cells:
        if cell.column <= COL_QTY_FACT:
            cell.fill = fill_to_apply

    wb.save(FILE_NAME)

    status = "🟢 ВСЕ ПРИШЛО" if new_fact >= qty_plan else "🟡 ЧАСТИЧНО"
    return f"Записано! {status}\nСтало: {new_fact} / {qty_plan}"


def get_warehouse_analytics():
    """Сбор статистики для Дашборда"""
    wb, ws = load_data()

    stats = {
        'total_positions': 0,
        'completed_positions': 0,
        'partial_positions': 0,
        'sum_plan': 0,
        'sum_fact': 0,
        'categories': {}
    }

    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            qty_plan = float(row[COL_QTY_PLAN - 1].value or 0)
            sum_plan = float(row[COL_SUM_PLAN - 1].value or 0)
            category = str(row[COL_CATEGORY - 1].value or "Прочее")[:15]

            qty_fact = float(ws.cell(row=idx, column=COL_QTY_FACT).value or 0)
            is_green_visually = is_row_green(row)

            if qty_fact == 0 and is_green_visually:
                qty_fact = qty_plan

            price_per_unit = sum_plan / qty_plan if qty_plan > 0 else 0
            sum_fact = qty_fact * price_per_unit
            if sum_fact > sum_plan: sum_fact = sum_plan

            stats['total_positions'] += 1
            stats['sum_plan'] += sum_plan
            stats['sum_fact'] += sum_fact

            if qty_fact >= qty_plan:
                stats['completed_positions'] += 1
            elif qty_fact > 0:
                stats['partial_positions'] += 1

            if category not in stats['categories']:
                stats['categories'][category] = {'plan': 0, 'fact': 0}
            stats['categories'][category]['plan'] += sum_plan
            stats['categories'][category]['fact'] += sum_fact

        except Exception:
            continue

    return stats


def get_full_database_file():
    return open(FILE_NAME, 'rb')