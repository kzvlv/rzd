import os
from datetime import date, datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response
# ИМПОРТЫ FLASK_LOGIN
from flask_login import login_required, current_user, login_user, logout_user
from functools import wraps
from datetime import date, datetime
from sqlalchemy.exc import IntegrityError
import csv
import io
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# 1. Импортируем "пустые" объекты из extensions.py
from extensions import db, bcrypt, login_manager

# 2. НЕ ИМПОРТИРУЕМ 'models' ЗДЕСЬ. Мы будем делать это внутри функций.

# --- Конфигурация Приложения ---
app = Flask(__name__)
app.config['SECRET_KEY'] = 'your_very_secret_key_change_this_asap'
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'database.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# 3. "Привязываем" объекты к нашему приложению
db.init_app(app)
bcrypt.init_app(app)
login_manager.init_app(app)

# --- ИСПРАВЛЕНИЕ ОШИБКИ: ДОБАВЛЯЕМ ФИЛЬТР DATE ---
@app.template_filter('date')
def date_filter(value, format='%Y-%m-%d'):
    if value == 'now':
        return datetime.now().strftime(format)
    if hasattr(value, 'strftime'):
        return value.strftime(format)
    return value

# --- Настройки Login Manager ---
login_manager.login_view = 'login'
login_manager.login_message_category = 'info'
login_manager.login_message = 'Пожалуйста, авторизуйтесь, чтобы получить доступ к этой странице.'


# --- Модели и Загрузчик ---
@login_manager.user_loader
def load_user(user_id):
    from models import User  # <-- Импорт ВНУТРИ функции
    return User.query.get(int(user_id))


# --- Декораторы для Ролей ---
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role != 'Admin':
            flash('У вас нет прав для доступа к этой странице.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)

    return decorated_function


def commandant_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.role not in ['Admin', 'Commandant']:
            flash('У вас нет прав для доступа к этой странице.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)

    return decorated_function


# --- Основные Маршруты (Страницы) ---

@app.route('/')
def index():
    from models import Room, Booking, Dorm
    # Добавляем импорт даты, если его нет
    from datetime import date

    today = date.today()  # <-- Получаем текущую дату

    try:
        total_capacity = db.session.query(db.func.sum(Room.capacity)).scalar()
        # Считаем занятые ТОЛЬКО на сегодня
        total_booked = Booking.query.filter(
            Booking.status.in_(['booked', 'living']),
            Booking.start_date <= today,
            Booking.end_date > today  # <-- Важное изменение для логики отелей (см. пункт 3)
        ).count()
        dorms_count = Dorm.query.count()
    except Exception as e:
        total_capacity = 0
        total_booked = 0
        dorms_count = 0

    stats = {
        "total_capacity": total_capacity or 0,
        "total_booked": total_booked,
        "total_free": (total_capacity or 0) - total_booked,
        "dorms_count": dorms_count,
        "today_date": today.strftime('%d.%m.%Y')  # <-- Передаем дату в шаблон
    }
    return render_template('index.html', stats=stats)


@app.route('/login', methods=['GET', 'POST'])
def login():
    from models import User  # <-- Импорт ВНУТРИ функции
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()

        if user and user.check_password(password):
            login_user(user, remember=True)
            flash(f'Добро пожаловать, {user.full_name}!', 'success')

            if user.role == 'Admin':
                return redirect(url_for('admin_panel'))
            elif user.role == 'Commandant':
                return redirect(url_for('commandant_panel'))
            else:
                return redirect(url_for('index'))
        else:
            flash('Неверный логин или пароль. Попробуйте снова.', 'danger')

    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Вы вышли из системы.', 'success')
    return redirect(url_for('login'))


@app.route('/booking')
@login_required
def booking():
    from models import Dorm  # <-- Импорт ВНУТРИ функции
    if current_user.role not in ['Admin', 'Enterprise']:
        flash('Только "Предприятия" и "Администратор" могут бронировать места.', 'danger')
        return redirect(url_for('index'))
    dorms = Dorm.query.all()
    return render_template('booking.html', dorms=dorms)


@app.route('/dorms')
def dorms_page():
    from models import Dorm  # <-- Импорт ВНУТРИ функции
    dorms = Dorm.query.all()
    return render_template('dorms.html', dorms=dorms)


@app.route('/rules')
def rules_page():
    return render_template('rules.html')


# --- Панели Управления ---

@app.route('/admin-panel')
@admin_required
def admin_panel():
    from models import User, Dorm, Room, Booking  # <-- Импорт ВНУТРИ функции
    today = date.today()
    enterprises = User.query.filter_by(role='Enterprise').all()
    commandants = User.query.filter_by(role='Commandant').all()
    dorms = Dorm.query.all()
    rooms = Room.query.all()

    # Разделяем брони
    future_bookings = Booking.query.filter(
        Booking.start_date > today,
        Booking.status == 'booked'
    ).order_by(Booking.start_date.asc()).all()

    # Изменяем логику: показываем только тех, кто СЕЙЧАС живет (активные)
    # То есть дата начала уже наступила, а дата выезда еще не прошла.
    past_bookings = Booking.query.filter(
        Booking.start_date <= today,
        Booking.end_date >= today  # <--- ЭТОТ ФИЛЬТР УБИРАЕТ "ПРОСРОЧЕННЫХ"
    ).order_by(Booking.start_date.desc()).all()

    return render_template('admin_panel.html',
                           enterprises=enterprises,
                           commandants=commandants,
                           rooms=rooms,
                           future_bookings=future_bookings,
                           past_bookings=past_bookings,
                           dorms=dorms)


@app.route('/admin/download-excel')
@admin_required
def download_excel_report():
    from models import Booking, Room, Dorm

    # 1. Получаем параметры из формы
    start_str = request.args.get('start_date')
    end_str = request.args.get('end_date')
    only_active = request.args.get('only_active')  # чекбокс

    if not start_str or not end_str:
        flash('Пожалуйста, выберите даты для отчета.', 'danger')
        return redirect(url_for('admin_panel'))

    try:
        filter_start = datetime.strptime(start_str, '%Y-%m-%d').date()
        filter_end = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        flash('Неверный формат дат.', 'danger')
        return redirect(url_for('admin_panel'))

    # 2. Формируем запрос
    query = Booking.query.join(Room).join(Dorm)

    # Фильтр пересечения дат: (StartA <= EndB) and (EndA >= StartB)
    # Находим всех, кто жил или будет жить в этот период
    query = query.filter(
        Booking.start_date <= filter_end,
        Booking.end_date >= filter_start
    )

    # Если галочка "Только активные" нажата, исключаем отмененные
    if only_active:
        query = query.filter(Booking.status.in_(['booked', 'living']))

    # Сортируем: Общежитие -> Комната -> Место
    bookings = query.order_by(Dorm.name, Room.room_number, Booking.bed_id).all()

    # 3. Создаем Excel файл через OpenPyXL
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Список проживающих"

    # Заголовки
    headers = [
        'Общежитие', 'Комната', 'Место',
        'ФИО', 'Пол', 'Группа', 'Предприятие',
        'Дата заезда', 'Дата выезда', 'Статус'
    ]
    ws.append(headers)

    # Стилизация заголовка
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Заполнение данными
    for b in bookings:
        status_ru = {
            'booked': 'Бронь',
            'living': 'Проживает',
            'completed': 'Выехал',
            'cancelled': 'Отмена'
        }.get(b.status, b.status)

        row = [
            b.room.dorm.name,
            b.room.room_number,
            b.bed_id,
            b.full_name,
            b.gender,
            b.study_group,
            b.enterprise_user.enterprise_name,
            b.start_date,  # Excel сам отформатирует дату
            b.end_date,
            status_ru
        ]
        ws.append(row)

    # Автоматическая ширина колонок
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 5

    # 4. Сохранение в память и отправка
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"report_{start_str}_{end_str}.xlsx"

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment;filename={filename}"}
    )

@app.route('/commandant-panel')
@commandant_required
def commandant_panel():
    from models import Booking, Dorm  # <-- Импорт ВНУТРИ функции
    today = date.today()

    # Администратор может видеть любую панель
    if current_user.role == 'Admin':
        dorm_id = request.args.get('dorm_id', type=int)
        if dorm_id:
            dorm = Dorm.query.get_or_404(dorm_id)
        else:
            dorm = Dorm.query.first()
            if not dorm:
                flash('В системе нет общежитий. Сначала добавьте их.', 'danger')
                return redirect(url_for('admin_panel'))
        all_dorms = Dorm.query.all()  # Админ может переключаться
    else:
        # Комендант видит только свое
        dorm = Dorm.query.get(current_user.dorm_id)
        if not dorm:
            flash('Ваш аккаунт не привязан к общежитию.', 'danger')
            return redirect(url_for('index'))
        all_dorms = None  # Комендант не может переключаться

    # Загружаем будущие заезды для таблицы
    room_ids = [room.id for room in dorm.rooms]
    future_bookings = Booking.query.filter(
        Booking.room_id.in_(room_ids),
        Booking.start_date > today,
        Booking.status == 'booked'
    ).order_by(Booking.start_date.asc()).all()

    return render_template('commandant_panel.html',
                           dorm=dorm,
                           all_dorms=all_dorms,
                           future_bookings=future_bookings)


@app.route('/enterprise-panel')
@login_required
def enterprise_panel():
    from models import Booking  # <-- Импорт ВНУТРИ функции
    if current_user.role != 'Enterprise':
        return redirect(url_for('index'))
    bookings = Booking.query.filter_by(enterprise_id=current_user.id).order_by(Booking.start_date.desc()).all()
    return render_template('enterprise_panel.html', bookings=bookings)


# --- НОВЫЙ БЛОК: ШАХМАТКА (TIMELINE) ---

@app.route('/timeline')
@commandant_required
def timeline():
    from models import Dorm
    if current_user.role == 'Admin':
        dorms = Dorm.query.all()
    else:
        # Если у коменданта нет общежития, берем первое попавшееся (или обработать ошибку)
        if not current_user.dorm_id:
            flash('Ваш аккаунт не привязан к общежитию.', 'danger')
            return redirect(url_for('index'))
        dorms = [Dorm.query.get(current_user.dorm_id)]
    return render_template('timeline.html', dorms=dorms)


@app.route('/api/timeline-data')
@login_required
def timeline_data():
    from models import Room, Booking, Dorm

    dorm_id = request.args.get('dorm_id', type=int)
    # Проверка прав
    if current_user.role == 'Commandant' and current_user.dorm_id != dorm_id:
        return jsonify({'error': 'Access denied'}), 403

    dorm = Dorm.query.get_or_404(dorm_id)

    groups = []
    items = []

    # 1. ГРУППЫ (Комнаты)
    for room in dorm.rooms:
        for i in range(1, room.capacity + 1):
            group_id = f"{room.id}_{i}"

            if room.status == 'repair':
                content_html = f"К.{room.room_number} <span class='text-red-500'>(Ремонт)</span>"
                # Фон ремонта
                items.append({
                    'id': f"repair_{group_id}",
                    'group': group_id,
                    'start': '2023-01-01',
                    'end': '2030-01-01',
                    'type': 'background',
                    'className': 'status-repair'
                })
            else:
                content_html = f"К.{room.room_number} <span class='text-gray-400 text-xs'>(М.{i})</span>"

            groups.append({'id': group_id, 'content': content_html})

    # 2. БРОНИРОВАНИЯ
    relevant_bookings = Booking.query.filter(
        Booking.room_id.in_([r.id for r in dorm.rooms]),
        Booking.status.in_(['booked', 'living'])
    ).all()

    for b in relevant_bookings:
        group_id = f"{b.room_id}_{b.bed_id}"

        # Определяем цвет
        css_class = 'status-living' if b.status == 'living' else 'status-booked'
        ent_name = "Неизвестно"
        if b.enterprise_user:
            ent_name = b.enterprise_user.enterprise_name or b.enterprise_user.full_name

        # Проверяем, есть ли ПРАКТИКА и попадает ли она в даты проживания
        has_practice = False
        if b.practice_start and b.practice_end:
            if b.start_date <= b.practice_start < b.end_date:
                has_practice = True

        if has_practice:
            # --- ВАРИАНТ С ПРАКТИКОЙ (РАЗРЫВ) ---

            # Часть 1: ДО ПРАКТИКИ
            if b.start_date < b.practice_start:
                items.append({
                    'id': f"{b.id}_part1",  # Уникальный ID части
                    'group': group_id,
                    'content': b.full_name,
                    'start': b.start_date.strftime('%Y-%m-%dT06:00:00'),
                    'end': b.practice_start.strftime('%Y-%m-%dT06:00:00'),
                    'className': css_class,
                    'title': f"{b.full_name} (До практики)",
                    'booking_id': b.id,
                    'enterprise': ent_name
                })

            # Часть 2: САМА ПРАКТИКА (Фон)
            items.append({
                'id': f"{b.id}_practice",
                'group': group_id,
                'content': '',
                'start': b.practice_start.strftime('%Y-%m-%dT06:00:00'),
                'end': b.practice_end.strftime('%Y-%m-%dT06:00:00'),
                'type': 'background',
                'className': 'status-practice-bg',  # Новый стиль
                'title': f"{b.full_name} убыл на практику",
                'booking_id': b.id  # Чтобы клик тоже работал
            })

            # Часть 3: ПОСЛЕ ПРАКТИКИ
            if b.practice_end < b.end_date:
                items.append({
                    'id': f"{b.id}_part2",
                    'group': group_id,
                    'content': b.full_name,
                    'start': b.practice_end.strftime('%Y-%m-%dT12:00:00'),
                    'end': b.end_date.strftime('%Y-%m-%dT12:00:00'),
                    'className': css_class,
                    'title': f"{b.full_name} (После практики)",
                    'booking_id': b.id,
                    'enterprise': ent_name
                })

        else:
            # --- ОБЫЧНЫЙ ВАРИАНТ (БЕЗ ПРАКТИКИ) ---
            items.append({
                'id': b.id,
                'group': group_id,
                'content': b.full_name,
                'start': b.start_date.strftime('%Y-%m-%dT12:00:00'),
                'end': b.end_date.strftime('%Y-%m-%dT12:00:00'),
                'className': css_class,
                'title': f"{b.full_name} ({ent_name})",
                'booking_id': b.id,
                'enterprise': ent_name
            })

    return jsonify({'groups': groups, 'items': items})


# --- НОВАЯ СТРАНИЦА: "КАРТА ЗАНЯТОСТИ" (ДЛЯ АДМИНА И КОМЕНДАНТА) ---
@app.route('/occupancy-map')
@commandant_required  # (Доступно коменданту и админу)
def occupancy_map():
    from models import Dorm
    # Админ видит все, комендант - только свое
    if current_user.role == 'Admin':
        dorms = Dorm.query.all()
    else:
        dorms = [Dorm.query.get(current_user.dorm_id)]

    return render_template('occupancy_map.html', dorms=dorms)


# --- API / Управление (Админ) ---

@app.route('/admin/create-user', methods=['POST'])
@admin_required
def create_user():
    from models import User  # <-- Импорт ВНУТРИ функции
    data = request.form
    try:
        if User.query.filter_by(username=data['username']).first():
            flash('Пользователь с таким логином уже существует.', 'danger')
            return redirect(url_for('admin_panel'))

        new_user = User(
            username=data['username'],
            full_name=data['full_name'],
            phone=data.get('phone'),
            role=data['role']
        )
        new_user.set_password(data['password'])

        if data['role'] == 'Enterprise':
            new_user.enterprise_name = data.get('enterprise_name')
        elif data['role'] == 'Commandant':
            new_user.dorm_id = int(data.get('dorm_id')) if data.get('dorm_id') else None

        db.session.add(new_user)
        db.session.commit()
        flash('Пользователь успешно создан.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при создании пользователя: {str(e)}', 'danger')
    return redirect(url_for('admin_panel'))


@app.route('/admin/delete-user/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    from models import User  # <-- Импорт ВНУТРИ функции
    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('Вы не можете удалить сами себя.', 'danger')
        return redirect(url_for('admin_panel'))

    try:
        db.session.delete(user)
        db.session.commit()
        flash(f'Пользователь {user.full_name} успешно удален.', 'success')
    except IntegrityError:
        db.session.rollback()
        flash(f'Невозможно удалить пользователя {user.full_name}, так как за ним закреплены бронирования.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Произошла ошибка при удалении: {str(e)}', 'danger')

    return redirect(url_for('admin_panel'))


@app.route('/admin/edit-user/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def edit_user(user_id):
    from models import User, Dorm  # <-- Импорт ВНУТРИ функции
    user = User.query.get_or_404(user_id)
    dorms = Dorm.query.all()

    if request.method == 'POST':
        data = request.form
        try:
            if user.username != data['username'] and User.query.filter_by(username=data['username']).first():
                flash('Этот логин уже занят.', 'danger')
                return render_template('edit_user.html', user=user, dorms=dorms)

            user.username = data['username']
            user.full_name = data['full_name']
            user.phone = data.get('phone')

            if user.role == 'Enterprise':
                user.enterprise_name = data.get('enterprise_name')
            elif user.role == 'Commandant':
                user.dorm_id = int(data.get('dorm_id')) if data.get('dorm_id') else None

            if data.get('password'):
                user.set_password(data['password'])

            db.session.commit()
            flash(f'Данные пользователя {user.full_name} успешно обновлены.', 'success')
            return redirect(url_for('admin_panel'))
        except Exception as e:
            db.session.rollback()
            flash(f'Ошибка при обновлении: {str(e)}', 'danger')

    return render_template('edit_user.html', user=user, dorms=dorms)


@app.route('/admin/create-room', methods=['POST'])
@admin_required
def create_room():
    from models import Room  # <-- Импорт ВНУТРИ функции
    data = request.form
    try:
        new_room = Room(
            dorm_id=data['dorm_id'],
            room_number=data['room_number'],
            capacity=data['capacity']
        )
        db.session.add(new_room)
        db.session.commit()
        flash(f'Комната {new_room.room_number} успешно добавлена.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при добавлении комнаты: {str(e)}', 'danger')
    return redirect(url_for('admin_panel'))


@app.route('/admin/delete-room/<int:room_id>', methods=['POST'])
@admin_required
def delete_room(room_id):
    from models import Room  # <-- Импорт ВНУТРИ функции
    room = Room.query.get_or_404(room_id)
    try:
        db.session.delete(room)
        db.session.commit()
        flash(f'Комната {room.room_number} успешно удалена.', 'success')
    except IntegrityError:
        db.session.rollback()
        flash(f'Невозможно удалить комнату {room.room_number}, так как за ней закреплены бронирования.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Произошла ошибка при удалении: {str(e)}', 'danger')
    return redirect(url_for('admin_panel'))


@app.route('/api/admin/room-status', methods=['POST'])
@admin_required
def update_room_status():
    from models import Room  # <-- Импорт ВНУТРИ функции
    data = request.json
    room = Room.query.get(data['room_id'])
    if not room:
        return jsonify({'success': False, 'message': 'Комната не найдена'}), 404

    room.status = data['status']
    db.session.commit()
    return jsonify({'success': True, 'message': 'Статус комнаты обновлен.'})


@app.route('/admin/download-report')
@admin_required
def download_report():
    from models import Booking, Room, Dorm, User  # <-- Импорт ВНУТРИ функции
    bookings = Booking.query.order_by(Booking.start_date).all()

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        'ID Брони', 'ФИО', 'Группа', 'Пол',
        'Общежитие', 'Комната', 'Место',
        'Дата Заезда', 'Дата Выезда',
        'Практика Начало', 'Практика Конец',
        'Предприятие', 'Статус'
    ])

    for b in bookings:
        writer.writerow([
            b.id, b.full_name, b.study_group, b.gender,
            b.room.dorm.name, b.room.room_number, b.bed_id,
            b.start_date, b.end_date,
            b.practice_start or '---', b.practice_end or '---',
            b.enterprise_user.enterprise_name, b.status
        ])

    output.seek(0)

    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment;filename=booking_report_{date.today()}.csv"}
    )


# --- БРОНИРОВАНИЕ (API ДЛЯ JS) ---

@app.route('/api/get-rooms')
@login_required
def get_rooms():
    from models import Dorm, Room, Booking  # <-- Импорт ВНУТРИ функции
    dorm_id = request.args.get('dorm_id')
    start_str = request.args.get('start_date')
    end_str = request.args.get('end_date')
    gender = request.args.get('gender')

    try:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    except:
        return jsonify({"error": "Неверный формат дат."}), 400

    dorm = Dorm.query.get(dorm_id)
    if not dorm:
        return jsonify({"error": "Общежитие не найдено"}), 404

    rooms_data = []
    for room in dorm.rooms:
        bed_data = {}

        if room.status == 'repair':
            bed_data = {i: {'status': 'repair'} for i in range(1, room.capacity + 1)}
        else:
            all_bookings = Booking.query.filter(
                Booking.room_id == room.id,
                Booking.status.in_(['booked', 'living']),
                Booking.start_date < end_date,  # <--- ИЗМЕНЕНИЕ ЗДЕСЬ
                Booking.end_date > start_date  # Здесь оставляем как было
            ).all()

            room_gender = None
            if all_bookings:
                room_gender = all_bookings[0].gender
            is_gender_ok = (room_gender is None) or (room_gender == gender)

            for i in range(1, room.capacity + 1):
                bed_status = 'available'
                occupant_name = ''

                if not is_gender_ok:
                    bed_status = 'unavailable'

                for booking in all_bookings:
                    if booking.bed_id == i:
                        is_on_practice = False
                        if booking.practice_start and booking.practice_end:
                            if booking.practice_start <= start_date and booking.practice_end >= end_date:
                                is_on_practice = True

                        if not is_on_practice:
                            if booking.enterprise_id == current_user.id:
                                bed_status = 'occupied_by_you'
                                occupant_name = booking.full_name
                            else:
                                bed_status = 'occupied'

                        break

                bed_data[i] = {'status': bed_status, 'occupant': occupant_name}

        beds_list = [{'id': bed_id, 'status': data['status'], 'occupant': data.get('occupant')} for bed_id, data in
                     bed_data.items()]
        rooms_data.append({'id': room.id, 'number': room.room_number, 'capacity': room.capacity, 'beds': beds_list})

    return jsonify(rooms_data)


@app.route('/api/book', methods=['POST'])
@login_required
def book_room():
    from models import Booking  # <-- Импорт ВНУТРИ функции
    if current_user.role not in ['Admin', 'Enterprise']:
        return jsonify({'success': False, 'message': 'Только Предприятия и Администратор могут бронировать.'}), 403

    data = request.json

    try:
        new_booking = Booking(
            room_id=data['roomId'],
            enterprise_id=current_user.id,
            full_name=data['fullName'],
            study_group=data['studyGroup'],
            gender=data['gender'],
            bed_id=data['bedId'],
            start_date=datetime.strptime(data['startDate'], '%Y-%m-%d').date(),
            end_date=datetime.strptime(data['endDate'], '%Y-%m-%d').date(),
            practice_start=datetime.strptime(data['practiceStart'], '%Y-%m-%d').date() if data.get(
                'practiceStart') else None,
            practice_end=datetime.strptime(data['practiceEnd'], '%Y-%m-%d').date() if data.get('practiceEnd') else None,
            status='booked'
        )
        db.session.add(new_booking)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Бронирование успешно!'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Ошибка: {str(e)}'}), 500


# --- УПРАВЛЕНИЕ БРОНИРОВАНИЕМ (Комендант / Предприятие) ---

@app.route('/booking/cancel/<int:booking_id>', methods=['POST'])
@login_required
def cancel_booking(booking_id):
    from models import Booking  # <-- Импорт ВНУТРИ функции
    booking = Booking.query.get_or_404(booking_id)

    if current_user.role == 'Enterprise' and booking.enterprise_id != current_user.id:
        flash('Вы не можете отменить чужое бронирование.', 'danger')
        return redirect(url_for('enterprise_panel'))

    try:
        booking.status = 'cancelled'
        db.session.commit()
        flash('Бронирование успешно отменено.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при отмене: {str(e)}', 'danger')

    if current_user.role == 'Enterprise':
        return redirect(url_for('enterprise_panel'))
    else:
        return redirect(url_for('admin_panel'))


@app.route('/booking/update-status/<int:booking_id>', methods=['POST'])
@commandant_required
def update_booking_status(booking_id):
    from models import Booking  # <-- Импорт ВНУТРИ функции
    booking = Booking.query.get_or_404(booking_id)

    if current_user.role == 'Commandant' and booking.room.dorm_id != current_user.dorm_id:
        flash('Вы не можете управлять бронированиями в чужом общежитии.', 'danger')
        return redirect(url_for('commandant_panel'))

    try:
        new_status = request.form.get('status')
        if new_status in ['booked', 'living', 'cancelled', 'completed']:
            booking.status = new_status
            db.session.commit()
            flash(f'Статус брони для {booking.full_name} обновлен.', 'success')
        else:
            flash('Неверный статус.', 'danger')
    except Exception as e:
        db.session.rollback()
        flash(f'Ошибка при обновлении статуса: {str(e)}', 'danger')

    if current_user.role == 'Admin':
        return redirect(url_for('admin_panel'))
    else:
        return redirect(url_for('commandant_panel'))


# --- ПЕРЕМЕЩЕНИЕ (ВИЗУАЛЬНОЕ) ---

@app.route('/move-booking/<int:booking_id>', methods=['GET'])
@commandant_required  # (Подходит и для Админа)
def move_booking_page(booking_id):
    from models import Booking, Dorm
    booking = Booking.query.get_or_404(booking_id)

    # Проверка прав
    if current_user.role == 'Commandant' and booking.room.dorm_id != current_user.dorm_id:
        flash('Вы можете перемещать гостей только в своем общежитии.', 'danger')
        return redirect(url_for('commandant_panel'))

    # Админ видит все общежития, Комендант - только свое
    if current_user.role == 'Admin':
        dorms = Dorm.query.all()
    else:
        dorms = [Dorm.query.get(current_user.dorm_id)]

    return render_template('move_booking.html', booking=booking, dorms=dorms)


@app.route('/move-booking/<int:booking_id>', methods=['POST'])
@commandant_required  # (Подходит и для Админа)
def move_booking_action(booking_id):
    from models import Booking, Room
    booking = Booking.query.get_or_404(booking_id)

    data = request.form
    new_room_id = data.get('new_room_id', type=int)
    new_bed_id = data.get('new_bed_id', type=int)

    if not new_room_id or not new_bed_id:
        flash('Ошибка: не получены ID комнаты или места.', 'danger')
        return redirect(url_for('move_booking_page', booking_id=booking_id))

    new_room = Room.query.get(new_room_id)

    # --- Валидация ---
    if current_user.role == 'Commandant' and new_room.dorm_id != current_user.dorm_id:
        flash('Вы не можете перемещать в чужое общежитие.', 'danger')
        return redirect(url_for('move_booking_page', booking_id=booking_id))

    is_occupied = Booking.query.filter(
        Booking.room_id == new_room_id,
        Booking.bed_id == new_bed_id,
        Booking.status.in_(['booked', 'living']),
        Booking.start_date < booking.end_date,  # <--- ИЗМЕНЕНИЕ ЗДЕСЬ
        Booking.end_date > booking.start_date,  # Здесь оставляем как было
        Booking.id != booking_id
    ).first()

    if is_occupied:
        flash('Ошибка: Новое место занято в эти даты.', 'danger')
        return redirect(url_for('move_booking_page', booking_id=booking_id))

    other_occupants = Booking.query.filter(
        Booking.room_id == new_room_id,
        Booking.id != booking_id,
        Booking.status.in_(['booked', 'living']),
        Booking.start_date <= booking.end_date,
        Booking.end_date >= booking.start_date
    ).all()

    if other_occupants:
        new_room_gender = other_occupants[0].gender
        if new_room_gender != booking.gender:
            flash(f'Ошибка: Нельзя заселить {booking.gender} в комнату, где проживают {new_room_gender}.', 'danger')
            return redirect(url_for('move_booking_page', booking_id=booking_id))

    # --- Перемещение ---
    try:
        old_room_number = booking.room.room_number
        booking.room_id = new_room_id
        booking.bed_id = new_bed_id
        db.session.commit()
        flash(
            f'{booking.full_name} успешно перемещен из к. {old_room_number} в к. {new_room.room_number}, место {new_bed_id}.',
            'success')

        if current_user.role == 'Admin':
            return redirect(url_for('admin_panel'))
        else:
            return redirect(url_for('commandant_panel'))

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': f'Ошибка сервера: {str(e)}'}), 500


# --- API ДЛЯ КАРТ ---

@app.route('/api/dorm-state/<int:dorm_id>')
@commandant_required
def get_dorm_state(dorm_id):
    from models import Room, Booking, User  # <-- Добавлен User

    if current_user.role == 'Commandant' and current_user.dorm_id != dorm_id:
        return jsonify({"error": "Доступ запрещен"}), 403

    rooms = Room.query.filter_by(dorm_id=dorm_id).all()
    rooms_data = []
    today = date.today()

    for room in rooms:
        bed_data = {}

        if room.status == 'repair':
            bed_data = {i: {'status': 'repair'} for i in range(1, room.capacity + 1)}
        else:
            current_bookings = Booking.query.filter(
                Booking.room_id == room.id,
                Booking.status.in_(['booked', 'living']),
                Booking.start_date <= today,
                Booking.end_date >= today
            ).join(User, User.id == Booking.enterprise_id).all()  # Join с User

            for i in range(1, room.capacity + 1):
                bed_status = 'available'
                booking_info = None

                for booking in current_bookings:
                    if booking.bed_id == i:
                        is_on_practice = False
                        if booking.practice_start and booking.practice_end:
                            if booking.practice_start <= today and booking.practice_end >= today:
                                is_on_practice = True

                        if not is_on_practice:
                            bed_status = 'occupied-living' if booking.status == 'living' else 'occupied-booked'
                            booking_info = {
                                'id': booking.id,
                                'full_name': booking.full_name,
                                'gender': booking.gender,
                                'group': booking.study_group,
                                'enterprise': booking.enterprise_user.enterprise_name,
                                # Теперь enterprise_user доступен
                                'start': booking.start_date.strftime('%d.%m.%Y'),
                                'end': booking.end_date.strftime('%d.%m.%Y')
                            }

                        break

                bed_data[i] = {'status': bed_status, 'booking': booking_info}

        beds_list = [{'id': bed_id, 'status': data['status'], 'booking': data['booking']} for bed_id, data in
                     bed_data.items()]
        rooms_data.append({'id': room.id, 'number': room.room_number, 'capacity': room.capacity, 'beds': beds_list})

    return jsonify(rooms_data)


# НОВЫЙ API для Карты Занятости
@app.route('/api/map-for-dates')
@commandant_required
def api_map_for_dates():
    from models import Dorm, Room, Booking, User  # <-- Добавлен User

    dorm_id = request.args.get('dorm_id', type=int)
    start_str = request.args.get('start_date')
    end_str = request.args.get('end_date')

    # Валидация
    if not dorm_id or not start_str or not end_str:
        return jsonify({"error": "Необходимы ID общежития и обе даты."}), 400

    try:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    except:
        return jsonify({"error": "Неверный формат дат."}), 400

    if current_user.role == 'Commandant' and current_user.dorm_id != dorm_id:
        return jsonify({"error": "Доступ запрещен"}), 403

    rooms = Room.query.filter_by(dorm_id=dorm_id).all()
    rooms_data = []

    for room in rooms:
        bed_data = {}
        if room.status == 'repair':
            bed_data = {i: {'status': 'repair'} for i in range(1, room.capacity + 1)}
        else:
            overlapping_bookings = Booking.query.filter(
                Booking.room_id == room.id,
                Booking.status.in_(['booked', 'living']),
                Booking.start_date < end_date,  # <--- ИЗМЕНЕНИЕ ЗДЕСЬ
                Booking.end_date > start_date
            ).join(User, User.id == Booking.enterprise_id).all()

            for i in range(1, room.capacity + 1):
                bed_status = 'available'
                booking_info = None

                for booking in overlapping_bookings:
                    if booking.bed_id == i:
                        is_on_practice = False
                        if booking.practice_start and booking.practice_end:
                            if booking.practice_start <= start_date and booking.practice_end >= end_date:
                                is_on_practice = True

                        if not is_on_practice:
                            # Логика: если даты брони ПОКРЫВАЮТ весь запрос - красный
                            # Если бронь только ЧАСТЬ запроса - оранжевый
                            if booking.start_date <= start_date and booking.end_date >= end_date:
                                bed_status = 'occupied'  # Полностью занято (Красный)
                            else:
                                bed_status = 'partial'  # Частично занято (Оранжевый)

                            booking_info = {
                                'id': booking.id,
                                'full_name': booking.full_name,
                                'group': booking.study_group,
                                'enterprise': booking.enterprise_user.enterprise_name,
                                'start': booking.start_date.strftime('%d.%m.%Y'),
                                'end': booking.end_date.strftime('%d.%m.%Y')
                            }
                        break

                bed_data[i] = {'status': bed_status, 'booking': booking_info}

        # ИСПРАВЛЕНИЕ: используем data.get('booking'), чтобы не было ошибки, если ключа нет (как при ремонте)
        beds_list = [{'id': bed_id, 'status': data['status'], 'booking': data.get('booking')} for bed_id, data
                     in
                     bed_data.items()]
        rooms_data.append({'id': room.id, 'number': room.room_number, 'capacity': room.capacity, 'beds': beds_list})

    return jsonify(rooms_data)


# --- НОВЫЕ API ДЛЯ ОЧЕРЕДИ ЗАСЕЛЕНИЯ ---

@app.route('/api/pending-arrivals')
@commandant_required
def get_pending_arrivals():
    from models import Booking, Room, Dorm
    dorm_id = request.args.get('dorm_id', type=int)

    # Проверка прав
    if current_user.role == 'Commandant' and current_user.dorm_id != dorm_id:
        return jsonify({'error': 'Access denied'}), 403

    today = date.today()

    # Ищем брони: Статус = booked И Дата заезда <= Сегодня
    # То есть те, кто должен заехать сегодня или уже опоздал
    pending_bookings = Booking.query.join(Room).filter(
        Room.dorm_id == dorm_id,
        Booking.status == 'booked',
        Booking.start_date <= today
    ).order_by(Booking.start_date.asc()).all()

    result = []
    for b in pending_bookings:
        # Определяем: это сегодня или просрочено?
        is_overdue = b.start_date < today

        result.append({
            'id': b.id,
            'full_name': b.full_name,
            'group': b.study_group,
            'room': b.room.room_number,
            'bed': b.bed_id,
            'start_date': b.start_date.strftime('%d.%m.%Y'),
            'end_date': b.end_date.strftime('%d.%m.%Y'),
            'is_overdue': is_overdue,
            'days_overdue': (today - b.start_date).days if is_overdue else 0,
            'enterprise': b.enterprise_user.enterprise_name
        })

    return jsonify(result)


@app.route('/api/quick-status-change', methods=['POST'])
@commandant_required
def quick_status_change():
    from models import Booking
    data = request.json
    booking_id = data.get('booking_id')
    action = data.get('action')  # 'checkin' (заселить) или 'cancel' (отменить)

    booking = Booking.query.get_or_404(booking_id)

    # Проверка прав коменданта
    if current_user.role == 'Commandant' and booking.room.dorm_id != current_user.dorm_id:
        return jsonify({'success': False, 'message': 'Нет доступа к этому общежитию'}), 403

    try:
        if action == 'checkin':
            booking.status = 'living'
            msg = f'Гость {booking.full_name} успешно заселен!'
        elif action == 'cancel':
            booking.status = 'cancelled'
            msg = f'Бронь {booking.full_name} отменена (незаезд).'
        else:
            return jsonify({'success': False, 'message': 'Неверное действие'}), 400

        db.session.commit()
        return jsonify({'success': True, 'message': msg})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/privacy')
def privacy_page():
    return render_template('privacy.html')


@app.route('/api/available-beds')
def get_available_beds():
    from models import Room, Booking, Bed
    from sqlalchemy import and_, or_

    dorm_id = request.args.get('dorm_id', type=int)
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    gender = request.args.get('gender')

    if not all([dorm_id, start_date_str, end_date_str, gender]):
        return jsonify([])

    # Конвертируем даты
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify([])

    # 1. Находим все кровати в этом общежитии, подходящие по полу (или пустые комнаты)
    # Логика: комната должна быть либо пустой, либо занята тем же полом
    # (Упрощение: ищем просто все кровати в общежитии, фильтруем занятые)

    # Получаем комнаты общежития
    rooms = Room.query.filter_by(dorm_id=dorm_id).all()
    available_rooms = []

    for room in rooms:
        # Пропускаем ремонт
        if room.status == 'repair':
            continue

        # Проверка пола комнаты (если реализована логика "женская/мужская комната")
        # Если в комнате уже кто-то живет, проверяем пол соседей
        room_bookings = Booking.query.filter(
            Booking.room_id == room.id,
            Booking.status == 'living',
            Booking.end_date > start_date,  # Пересечение дат
            Booking.start_date < end_date
        ).all()

        # Если комната занята людьми другого пола - пропускаем
        current_genders = {b.gender for b in room_bookings if b.gender}
        if current_genders and gender not in current_genders:
            continue  # Комната занята другим полом

        # Находим свободные кровати в этой комнате
        beds_data = []
        for bed in room.beds:
            # Проверяем брони на эту конкретную кровать
            is_occupied = Booking.query.filter(
                Booking.bed_id == bed.id,
                Booking.status.in_(['booked', 'living']),
                # Логика пересечения интервалов:
                # (StartA < EndB) and (EndA > StartB)
                and_(Booking.start_date < end_date, Booking.end_date > start_date)
            ).first()

            if not is_occupied:
                beds_data.append({
                    'id': bed.id,
                    'bed_number': bed.bed_number
                })

        if beds_data:
            available_rooms.append({
                'id': room.id,
                'number': room.room_number,
                'beds': beds_data
            })

    return jsonify(available_rooms)
@app.route('/api/evict-early', methods=['POST'])
@commandant_required
def evict_early():
    from models import Booking

    data = request.json
    booking_id = data.get('booking_id')

    booking = Booking.query.get_or_404(booking_id)

    # Проверка прав (свое ли общежитие)
    if current_user.role == 'Commandant' and booking.room.dorm_id != current_user.dorm_id:
        return jsonify({'success': False, 'message': 'Ошибка доступа: чужое общежитие'}), 403

    try:
        # 1. Меняем статус на "Завершено" (уходит в историю)
        booking.status = 'completed'

        # 2. ВАЖНО: Обрезаем дату выезда до СЕГОДНЯШНЕГО дня.
        # Это освобождает место в календаре начиная с завтрашнего дня (или сейчас, в зависимости от логики).
        booking.end_date = date.today()

        db.session.commit()
        return jsonify({'success': True, 'message': f'Гость {booking.full_name} успешно выселен.'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500